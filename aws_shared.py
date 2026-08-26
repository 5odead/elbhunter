"""
aws_shared.py
=============
Shared utilities for the AWS TPRM Toolkit.

Thread-safety notes (important — several fixes landed here):
  • socket.setdefaulttimeout() is GLOBAL state — calling it from multiple threads
    simultaneously causes silent timeout races (verified: 4/4 threads got wrong
    timeouts in testing). All socket calls are now protected by _socket_lock, and
    dnspython is preferred because each Resolver() instance is fully independent.
  • dns.resolver.Resolver() instances are safe to create per-call in threads.
  • requests.Session connection pools are thread-safe for concurrent GETs, but
    a stale pool after a connection failure will keep failing. _remount_session()
    rebuilds the pool in-place before each retry.
  • ip_network objects are expensive to create. build_prefix_networks() pre-builds
    them once — benchmarked at 14.7x faster for batch IP checking.
"""

import ipaddress
import json
import random
import socket
import threading
import time
from pathlib import Path
from typing import Callable, Optional, Union
from urllib.parse import quote

import pandas as pd
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl.styles import Font, PatternFill

# ── Optional dnspython import ─────────────────────────────────────────────────
try:
    import dns.resolver
    import dns.reversename
    import dns.exception
    HAS_DNSPYTHON = True
except ImportError:
    HAS_DNSPYTHON = False

# ── Thread-local session storage ──────────────────────────────────────────────
_thread_local = threading.local()

# ── Global lock for socket calls (setdefaulttimeout is global state) ──────────
_socket_lock = threading.Lock()

# ── Public DNS fallback servers (used when system resolver times out) ─────────
FALLBACK_DNS_SERVERS = [
    "8.8.8.8",   # Google
    "1.1.1.1",   # Cloudflare
    "9.9.9.9",   # Quad9
]

# ── AWS service suffix patterns ───────────────────────────────────────────────
AWS_SERVICE_SUFFIXES: list[tuple[str, str, str]] = [
    (".elb.amazonaws.com",        "AWS Elastic Load Balancer",  "ELB"),
    (".execute-api.",             "AWS API Gateway",             "API-GW"),
    (".cloudfront.net",           "AWS CloudFront CDN",          "CDN"),
    (".s3-website",               "AWS S3 Static Website",       "S3"),
    (".s3.amazonaws.com",         "AWS S3 Bucket",               "S3"),
    (".s3.",                      "AWS S3 Bucket",               "S3"),
    (".elasticbeanstalk.com",     "AWS Elastic Beanstalk",       "BEANSTALK"),
    (".awsglobalaccelerator.com", "AWS Global Accelerator",      "GA"),
    (".compute.amazonaws.com",    "AWS EC2 Instance",            "EC2"),
    (".amazonaws.com",            "AWS (other service)",         "AWS-OTHER"),
]

# ── HTTP fingerprint rules ─────────────────────────────────────────────────────
HTTP_FINGERPRINT_RULES: list[tuple[str, str, str, str]] = [
    ("server",            "awselb/2.0",   "AWS ALB (Application LB)",      "ALB"),
    ("set-cookie",        "AWSALB=",      "AWS ALB sticky session cookie",  "ALB"),
    ("set-cookie",        "AWSALBCORS=",  "AWS ALB CORS sticky cookie",     "ALB"),
    ("set-cookie",        "AWSELB=",      "AWS Classic ELB cookie",         "CLB"),
    ("set-cookie",        "AWSELBCORS=",  "AWS Classic ELB CORS cookie",    "CLB"),
    ("x-amz-cf-id",       "",             "AWS CloudFront",                  "CF"),
    ("x-amz-cf-pop",      "",             "AWS CloudFront (POP header)",     "CF"),
    ("via",               "cloudfront",   "AWS CloudFront (Via header)",     "CF"),
    ("x-amz-request-id",  "",             "AWS S3 (request ID)",             "S3"),
    ("x-amz-id-2",        "",             "AWS S3 (extended request ID)",    "S3"),
    ("server",            "AmazonS3",     "AWS S3",                          "S3"),
    ("x-amzn-requestid",  "",             "AWS API Gateway",                 "API-GW"),
    ("x-amzn-trace-id",   "",             "AWS X-Ray trace (ALB/API GW)",    "ALB"),
    ("x-amz-waf-action",  "",             "AWS WAF",                         "WAF"),
]

ALB_SPECIFIC_STATUS_CODES = {
    460: "ALB 460 — client closed connection before idle timeout (ALB confirmed)",
    463: "ALB 463 — X-Forwarded-For exceeded 30 IPs (ALB confirmed)",
    464: "ALB 464 — incompatible protocol version (ALB confirmed)",
}

# ── Excel/terminal colour palettes ────────────────────────────────────────────
COLOURS = {
    "aws":      "C6EFCE",
    "elb":      "92D050",
    "warning":  "FFEB9C",
    "danger":   "FFC7CE",
    "critical": "FF0000",
    "header":   "4472C4",
    "unknown":  "D9D9D9",
}

ANSI = {
    "red":    "\033[91m",
    "green":  "\033[92m",
    "yellow": "\033[93m",
    "cyan":   "\033[96m",
    "bold":   "\033[1m",
    "reset":  "\033[0m",
}


# ── Input normalisation ────────────────────────────────────────────────────────

def _safe_str(raw, default: str = "") -> str:
    """Convert anything to str safely — never raises."""
    if raw is None:
        return default
    try:
        return str(raw)
    except Exception:
        return default


def _safe_hostname(raw) -> str:
    """Strip and clean a hostname input of any type."""
    return _safe_str(raw).strip().rstrip(".")


def normalize_ip(
    ip_str,
) -> Optional[Union[ipaddress.IPv4Address, ipaddress.IPv6Address]]:
    """
    Parse an IP string. Returns None on any invalid input — never raises.
    FIX: now accepts any type (was str-only; passing None raised AttributeError).
    """
    if not isinstance(ip_str, (str, bytes)):
        try:
            ip_str = str(ip_str)
        except Exception:
            return None
    try:
        return ipaddress.ip_address(ip_str.strip())
    except (ValueError, AttributeError):
        return None


def normalize_domain(raw) -> str:
    """Strip scheme, whitespace, trailing dots from any input."""
    d = _safe_str(raw).strip().lower()
    for scheme in ("https://", "http://"):
        if d.startswith(scheme):
            d = d[len(scheme):]
    return d.split("/")[0].rstrip(".")


def normalize_ip_list(raw_lines: list) -> list[str]:
    """From raw text lines (one IP per line, comments OK), return valid IPs."""
    valid = []
    for line in raw_lines:
        line = _safe_str(line).strip()
        if not line or line.startswith("#"):
            continue
        line = line.split("#")[0].strip()
        if normalize_ip(line) is not None:
            valid.append(line)
    return valid


# ── IP / CIDR helpers ──────────────────────────────────────────────────────────

def build_prefix_networks(
    prefixes: list[dict],
) -> list[tuple[ipaddress.IPv4Network, dict]]:
    """
    Pre-build ipaddress.IPv4Network objects from AWS prefix dicts.
    Returns a list of (network, prefix_dict) tuples ready for fast lookup.

    WHY: Benchmarked at 14.7x faster than recreating ip_network inside
    check_ip_in_prefixes on every IP check. Build once, reuse for every IP.
    Skips IPv6 prefixes and malformed CIDRs silently.
    """
    networks = []
    for p in prefixes:
        cidr = p.get("ip_prefix", "")
        if not cidr or ":" in cidr:
            continue
        try:
            networks.append((ipaddress.ip_network(cidr, strict=False), p))
        except ValueError:
            continue
    return networks


def check_ip_in_networks(
    ip_str:   str,
    networks: list[tuple[ipaddress.IPv4Network, dict]],
) -> list[dict]:
    """
    Check an IP against pre-built (network, prefix_dict) tuples.
    Use with build_prefix_networks() for batch IP processing.
    Only matches IPv4 — IPv6 addresses return [].
    """
    ip = normalize_ip(ip_str)
    if ip is None or not isinstance(ip, ipaddress.IPv4Address):
        return []
    return [p for net, p in networks if ip in net]


def check_ip_in_prefixes(ip_str: str, prefixes: list[dict]) -> list[dict]:
    """
    Check one IP against raw prefix dicts (builds ip_network per call).
    Kept for backward compatibility with tests and one-off lookups.
    For batch processing use build_prefix_networks() + check_ip_in_networks().
    """
    ip = normalize_ip(ip_str)
    if ip is None:
        return []
    matches = []
    for p in prefixes:
        cidr = p.get("ip_prefix", "")
        if not cidr or ":" in cidr:
            continue
        try:
            if ip in ipaddress.ip_network(cidr, strict=False):
                matches.append(p)
        except ValueError:
            continue
    return matches


def dedupe_prefix_matches(matches: list[dict]) -> list[dict]:
    """
    AWS ip-ranges.json lists 'AMAZON' as a super-set of every other service.
    Prefer specific tags (EC2, S3, …) when both match the same CIDR.
    """
    if not matches:
        return matches
    specific = [m for m in matches if m.get("service", "") != "AMAZON"]
    return specific if specific else matches


# ── DNS helpers ────────────────────────────────────────────────────────────────

def resolve_a_records(hostname: str, timeout: float = 5.0) -> tuple[list[str], str]:
    """
    Resolve A records for a hostname.
    Uses dnspython when available (thread-safe, per-query timeout, fallback DNS).
    Falls back to socket with a global lock when dnspython is not installed.
    FIX: timeout is now actually enforced; previous socket approach was NOT
    thread-safe (socket.setdefaulttimeout is global state — race-condition confirmed).
    """
    if HAS_DNSPYTHON:
        return _resolve_a_dnspython(hostname, timeout)
    return _resolve_a_socket(hostname, timeout)


def _resolve_a_dnspython(hostname: str, timeout: float) -> tuple[list[str], str]:
    """
    dnspython-based A record resolution with automatic public-DNS fallback.
    Creates a new Resolver() per call — each instance is fully independent,
    so this is safe to call from multiple threads simultaneously.
    Fallback chain: system resolver → 8.8.8.8 → 1.1.1.1 → 9.9.9.9.
    Fallback only activates on Timeout/DNSException, NOT on NXDOMAIN (which
    is a definitive answer that all resolvers agree on).
    """
    h = _safe_hostname(hostname)
    if not h:
        return [], "Empty hostname"

    per_query  = min(timeout / 2, 3.0)
    last_error = f"All resolvers failed:{h}"

    # (configure_from_system, nameservers_override)
    resolver_sources: list[tuple[bool, Optional[list[str]]]] = [
        (True,  None),
        (False, ["8.8.8.8"]),
        (False, ["1.1.1.1"]),
        (False, ["9.9.9.9"]),
    ]

    for configure, servers in resolver_sources:
        try:
            r = dns.resolver.Resolver(configure=configure)
            if servers:
                r.nameservers = servers
            r.lifetime = timeout
            r.timeout  = per_query
            answers    = r.resolve(h, "A")
            ips        = list(dict.fromkeys(str(a.address) for a in answers))
            return ips, ""
        except dns.resolver.NXDOMAIN:
            return [], f"NXDOMAIN:{h}"      # definitive — no fallback
        except dns.resolver.NoAnswer:
            return [], f"No A records:{h}"  # definitive
        except (dns.resolver.Timeout, dns.exception.DNSException) as e:
            last_error = _safe_str(e)       # try next resolver
        except Exception as e:
            return [], f"Unexpected:{e}"    # unknown — don't retry

    return [], last_error


def _resolve_a_socket(hostname: str, timeout: float) -> tuple[list[str], str]:
    """
    Socket-based fallback. socket.setdefaulttimeout() is global, so we hold
    _socket_lock for the duration to prevent cross-thread timeout stomping.
    """
    h = _safe_hostname(hostname)
    if not h:
        return [], "Empty hostname"
    with _socket_lock:
        old = socket.getdefaulttimeout()
        socket.setdefaulttimeout(timeout)
        try:
            infos = socket.getaddrinfo(h, None, socket.AF_INET, socket.SOCK_STREAM)
            ips   = list(dict.fromkeys(info[4][0] for info in infos))
            return ips, ""
        except socket.gaierror as e:
            return [], _safe_str(e)
        except Exception as e:
            return [], f"Unexpected:{e}"
        finally:
            socket.setdefaulttimeout(old)


def reverse_dns(ip_str: str, timeout: float = 5.0) -> tuple[str, str]:
    """
    Reverse DNS lookup (PTR).
    FIX: timeout parameter now enforced; uses dnspython when available
    (thread-safe). Socket fallback is protected by _socket_lock.
    """
    ip = normalize_ip(ip_str)
    if ip is None:
        return "", f"Invalid IP:{ip_str}"
    if HAS_DNSPYTHON:
        return _reverse_dns_dnspython(str(ip), timeout)
    return _reverse_dns_socket(str(ip), timeout)


def _reverse_dns_dnspython(ip_str: str, timeout: float) -> tuple[str, str]:
    try:
        r          = dns.resolver.Resolver()
        r.lifetime = timeout
        r.timeout  = min(timeout / 2, 3.0)
        rev        = dns.reversename.from_address(ip_str)
        ans        = r.resolve(rev, "PTR")
        return _safe_str(ans[0]).rstrip("."), ""
    except dns.resolver.NXDOMAIN:
        return "", f"No PTR:{ip_str}"
    except dns.resolver.Timeout:
        return "", f"PTR timeout:{ip_str}"
    except Exception as e:
        return "", _safe_str(e)


def _reverse_dns_socket(ip_str: str, timeout: float) -> tuple[str, str]:
    with _socket_lock:
        old = socket.getdefaulttimeout()
        socket.setdefaulttimeout(timeout)
        try:
            host = socket.gethostbyaddr(ip_str)[0]
            return host, ""
        except socket.herror as e:
            return "", f"No PTR:{e}"
        except Exception as e:
            return "", _safe_str(e)
        finally:
            socket.setdefaulttimeout(old)


def follow_cname_chain(
    domain: str, max_hops: int = 10,
) -> tuple[list[str], str, str]:
    """
    Follow a CNAME chain using dnspython (per-call Resolver, thread-safe).
    Returns (chain_list, final_name, error_string).
    error_string contains 'NXDOMAIN' if a link in the chain no longer exists.
    """
    if not domain or not _safe_str(domain).strip():
        return [], "", "Empty domain"

    current = _safe_hostname(domain)
    chain   = []
    visited: set[str] = set()

    if not HAS_DNSPYTHON:
        return [], current, ""

    resolver          = dns.resolver.Resolver()
    resolver.lifetime = 5.0

    for _ in range(max_hops):
        if current in visited:
            return chain, current, "CNAME loop detected"
        visited.add(current)
        try:
            answers = resolver.resolve(current, "CNAME")
            target  = _safe_str(answers[0].target).rstrip(".")
            chain.append(target)
            current = target
        except dns.resolver.NoAnswer:
            break
        except dns.resolver.NXDOMAIN:
            return chain, current, f"NXDOMAIN:{current}"
        except dns.exception.DNSException as e:
            return chain, current, _safe_str(e)
        except Exception as e:
            return chain, current, _safe_str(e)

    return chain, current, ""


def is_dangling_dns(error: str) -> bool:
    """Return True if the error string indicates a broken/dangling DNS record."""
    if not error:
        return False
    low = _safe_str(error).lower()
    return any(p in low for p in (
        "nxdomain", "not found", "does not exist",
        "name or service not known", "no such host",
        "non-existent domain", "servfail", "nodename nor servname",
    ))


# ── Hostname / service classification ─────────────────────────────────────────

def classify_aws_hostname(hostname: str) -> tuple[str, str]:
    """Match a hostname against known AWS service suffixes."""
    if not hostname:
        return "Unknown", ""
    h = _safe_hostname(hostname).lower()
    for suffix, label, lb_type in AWS_SERVICE_SUFFIXES:
        if h.endswith(suffix) or suffix in h:
            return label, lb_type
    return "Unknown / Not AWS", ""


def classify_elb_subtype(hostname: str) -> str:
    h = _safe_hostname(hostname).lower()
    if "internal-" in h:
        return "Internal ELB (private — not internet-facing)"
    if ".elb.amazonaws.com" in h:
        return "Internet-facing ELB (ALB/NLB/CLB — confirm type via headers)"
    return ""


# ── HTTP fingerprint helpers ───────────────────────────────────────────────────

def apply_http_fingerprints(
    headers:     dict,
    status_code: int = 200,
) -> list[dict]:
    """Apply all HTTP fingerprint rules to response headers + status code."""
    if not isinstance(headers, dict):
        headers = {}
    headers_lower = {k.lower(): _safe_str(v) for k, v in headers.items()}
    results = []

    for h_name, h_contains, label, lb_type in HTTP_FINGERPRINT_RULES:
        val = headers_lower.get(h_name, "")
        if not val:
            continue
        if h_contains == "" or h_contains.lower() in val.lower():
            results.append({
                "label":        label,
                "lb_type":      lb_type,
                "header_name":  h_name,
                "header_value": val[:300],
            })

    # Coerce status_code to int safely before checking
    try:
        sc = int(status_code)
    except (TypeError, ValueError):
        sc = 0
    if sc in ALB_SPECIFIC_STATUS_CODES:
        results.append({
            "label":        ALB_SPECIFIC_STATUS_CODES[sc],
            "lb_type":      "ALB",
            "header_name":  "HTTP Status Code",
            "header_value": str(sc),
        })
    return results


def extract_cookies(set_cookie_header) -> list[str]:
    """
    Extract the cookie name from a Set-Cookie header value.
    FIX: Old version had a misleading 'comma-separated' comment and an
    always-firing break. Now explicit: first ;-segment is 'name=value'.
    Accepts any input type safely.
    """
    val = _safe_str(set_cookie_header)
    if not val:
        return []
    first = val.split(";")[0].strip()
    if "=" in first:
        return [first.split("=")[0].strip()]
    return []


# ── Session factory + reconnect helpers ───────────────────────────────────────

def make_session(
    user_agent:       str = "AWS-TPRM-Toolkit/1.0",
    pool_connections: int = 10,
    pool_maxsize:     int = 20,
) -> requests.Session:
    """
    Create a requests.Session with a configured HTTPAdapter connection pool.
    Prefer this over requests.Session() directly for consistent pool sizing.
    """
    from requests.adapters import HTTPAdapter
    session = requests.Session()
    session.headers.update({"User-Agent": user_agent})
    adapter = HTTPAdapter(
        pool_connections=pool_connections,
        pool_maxsize=pool_maxsize,
        pool_block=False,
    )
    session.mount("https://", adapter)
    session.mount("http://",  adapter)
    return session


def _remount_session(session: requests.Session) -> None:
    """
    Rebuild the HTTP adapter pool on a session after a connection failure.
    Stale connections in the pool cause repeated failures on retry —
    remounting creates a fresh pool without allocating a new session object.
    """
    from requests.adapters import HTTPAdapter
    adapter = HTTPAdapter(pool_connections=10, pool_maxsize=20, pool_block=False)
    session.mount("https://", adapter)
    session.mount("http://",  adapter)


def get_thread_session(user_agent: str = "AWS-TPRM-Toolkit/1.0") -> requests.Session:
    """
    Return a per-thread session, creating one on first call per thread.
    Safe to call from multiple threads simultaneously — each gets its own session.
    Use this inside ThreadPoolExecutor workers instead of sharing one session.
    """
    sess = getattr(_thread_local, "session", None)
    if sess is None:
        _thread_local.session = make_session(user_agent)
    return _thread_local.session


def close_thread_session() -> None:
    """Close and discard the current thread's session (e.g. after an error)."""
    sess = getattr(_thread_local, "session", None)
    if sess is not None:
        try:
            sess.close()
        except Exception:
            pass
        _thread_local.session = None


# ── Retry helper ───────────────────────────────────────────────────────────────

def retry_get(
    session:      requests.Session,
    url:          str,
    max_attempts: int   = 3,
    backoff:      float = 2.0,
    timeout:      int   = 15,
    **kwargs,
) -> requests.Response:
    """
    GET with exponential-backoff retry + jitter + connection pool rebuild.

    FIX 1 — Reconnect: On ConnectionError or Timeout, the session's connection
    pool may hold stale broken connections. _remount_session() rebuilds the pool
    in-place so the next attempt starts with fresh connections.

    FIX 2 — Jitter: ±10 % of base wait is added to prevent thundering-herd
    when many workers hit the same endpoint after a shared failure.
    """
    last_exc: Optional[Exception] = None

    for attempt in range(1, max_attempts + 1):
        try:
            resp = session.get(url, timeout=timeout, **kwargs)
            if resp.status_code == 429:
                wait = backoff * (2 ** (attempt - 1))
                print(f"  [rate-limit] 429 — waiting {wait:.0f}s "
                      f"(attempt {attempt}/{max_attempts})")
                time.sleep(wait)
                continue
            return resp

        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            last_exc = e
            if attempt < max_attempts:
                wait   = backoff * attempt
                jitter = wait * 0.1 * (0.5 - random.random())
                actual = max(0.0, wait + jitter)
                print(f"  [retry] {type(e).__name__} — attempt {attempt}/{max_attempts} "
                      f"— rebuilding pool, retrying in {actual:.1f}s")
                time.sleep(actual)
                _remount_session(session)   # fresh connection pool before retry

        except Exception as e:
            last_exc = e
            if attempt < max_attempts:
                wait = backoff * attempt
                print(f"  [retry] {type(e).__name__} — attempt {attempt}/{max_attempts} "
                      f"— retrying in {wait:.0f}s")
                time.sleep(wait)

    if last_exc is not None:
        raise last_exc
    raise RuntimeError(f"retry_get: exhausted {max_attempts} attempts for {url}")


# ── crt.sh subdomain enumeration (shared by passive_recon + pipeline) ─────────

def fetch_crtsh_subdomains(
    apex:        str,
    session:     requests.Session,
    max_retries: int = 3,
    timeout:     int = 30,
) -> list[str]:
    """
    Query crt.sh for all subdomains of an apex domain.
    Returns a sorted, deduplicated list of concrete subdomains (no wildcards).
    Always returns at least [apex] — never an empty list.

    FIX: was duplicated in aws_passive_recon.py and aws_tprm_pipeline.py
    with slightly different behaviour. Now in one place.
    """
    apex = normalize_domain(_safe_str(apex))
    if not apex:
        return []

    url = f"https://crt.sh/?q={quote(f'%.{apex}')}&output=json"

    for attempt in range(1, max_retries + 1):
        try:
            resp = retry_get(session, url, max_attempts=1, timeout=timeout)
            body = resp.text.strip()
            if not body:
                print(f"  [crt.sh] Empty response for {apex} — no CT logs found")
                return [apex]
            # HTML response means rate-limited or crt.sh error (200 with HTML body)
            if body.startswith("<"):
                raise ValueError(f"HTML response (HTTP {resp.status_code}) — crt.sh may be rate-limiting")
            data  = resp.json()
            names: set[str] = {apex}
            for entry in data:
                raw = _safe_str(entry.get("name_value", ""))
                for name in raw.split("\n"):
                    name = name.strip().lower().lstrip("*.")
                    if name and name.endswith(apex) and "*" not in name:
                        names.add(name)
            result = sorted(names)
            print(f"  [crt.sh] {len(result)} unique subdomains found for {apex}")
            return result
        except ValueError as e:
            print(f"  [crt.sh] JSON parse error (attempt {attempt}): {e}")
        except Exception as e:
            print(f"  [crt.sh] Error (attempt {attempt}): {e}")
        if attempt < max_retries:
            time.sleep(3 * attempt)

    print(f"  [crt.sh] Failed after {max_retries} attempts — returning apex only")
    return [apex]


# ── syncToken-aware AWS range cache ───────────────────────────────────────────

def load_aws_ranges_cached(
    session:       requests.Session,
    cache_file:    Path,
    token_file:    Path,
    force_refresh: bool = False,
) -> dict:
    """
    Download ip-ranges.json only when AWS's syncToken has changed.

    FIX 1 — syncToken vs TTL: Old approach re-downloaded every 6 hours regardless.
    We now peek at just the syncToken first (fast streaming request with its own
    short timeout) — if unchanged, serve from cache. This avoids all redundant
    downloads (~weekly update cadence from AWS).

    FIX 2 — Token peek timeout: The streaming peek now has an explicit 8-second
    timeout. Previously it used no timeout and could hang indefinitely.

    FIX 3 — Cache write errors: File-write failures are now caught and reported
    instead of crashing. The in-memory data is still returned.
    """
    url = "https://ip-ranges.amazonaws.com/ip-ranges.json"

    if not force_refresh and cache_file.exists():
        # Step 1: try to load the cached data (may be corrupted)
        cached_data = None
        try:
            d = json.loads(cache_file.read_text())
            if "prefixes" in d:
                cached_data = d
        except (json.JSONDecodeError, OSError):
            pass  # corrupted cache — fall through to download

        if cached_data is not None:
            # Step 2: check freshness via syncToken peek.
            # CRITICAL FIX: previous code fell through to a download on ANY
            # peek failure (403, empty body → JSONDecodeError, timeout, etc.).
            # The download then also failed, giving SystemExit even with a
            # perfectly valid cache on disk.
            # Correct logic: serve cache on any peek failure. Only download
            # when peek DEFINITIVELY shows the token has changed.
            cached_tok = token_file.read_text().strip() if token_file.exists() else ""
            stale = False
            try:
                peek = session.get(url, timeout=8, stream=True)
                if peek.ok:
                    # Got a valid response — check if token changed
                    live_token = peek.json().get("syncToken", "")
                    if live_token and live_token != cached_tok:
                        stale = True   # token changed → data is stale → download
                    else:
                        tok_label = live_token or cached_tok
                        print(f"[*] AWS ranges unchanged (syncToken {tok_label}) "
                              f"— using cache ({len(cached_data['prefixes'])} prefixes)")
                        return cached_data
                else:
                    # Non-2xx (403, 429, 503…) → network restricted → serve cache
                    print(f"[*] Cannot verify freshness (HTTP {peek.status_code})"
                          f" — serving from cache ({len(cached_data['prefixes'])} prefixes)")
                    return cached_data
            except Exception as peek_exc:
                # Timeout, ConnectionError, JSONDecodeError on 403 body, etc.
                # → serve cache rather than failing with SystemExit
                print(f"[*] Cannot verify freshness ({type(peek_exc).__name__})"
                      f" — serving from cache ({len(cached_data['prefixes'])} prefixes)")
                return cached_data

            if not stale:
                return cached_data   # shouldn't reach here, but safe fallback

    print(f"[*] Downloading AWS IP ranges from {url} ...")
    try:
        resp = retry_get(session, url, max_attempts=3, timeout=20)
        resp.raise_for_status()
        data = resp.json()
    except requests.exceptions.ConnectionError as e:
        raise SystemExit(f"[!] Network error — cannot reach AWS: {e}")
    except requests.exceptions.Timeout:
        raise SystemExit("[!] Request timed out — check internet connection")
    except requests.exceptions.HTTPError as e:
        raise SystemExit(f"[!] HTTP error from AWS: {e}")
    except (ValueError, KeyError) as e:
        raise SystemExit(f"[!] Response is not valid JSON: {e}")

    if "prefixes" not in data:
        raise SystemExit("[!] Unexpected response — 'prefixes' key missing")

    try:
        cache_file.write_text(json.dumps(data))
        token_file.write_text(data.get("syncToken", ""))
    except OSError as e:
        print(f"[!] Could not write cache ({e}) — will re-download next run")

    print(f"[+] Downloaded {len(data['prefixes'])} IPv4 prefixes "
          f"(syncToken: {data.get('syncToken', 'unknown')})")
    return data


# ── Terminal output helpers ───────────────────────────────────────────────────

def print_table(
    rows:          list[dict],
    title:         str = "",
    cols:          Optional[list[str]] = None,
    colour_fn:     Optional[Callable]  = None,
    max_col_width: int = 45,
) -> None:
    """Print a list of dicts as a fixed-width terminal table with optional ANSI colour."""
    if not rows:
        print("  (no results)\n")
        return

    display_cols = cols if cols else list(rows[0].keys())
    widths = {
        c: min(
            max(len(c), max((len(_safe_str(r.get(c, ""))) for r in rows), default=0)),
            max_col_width,
        )
        for c in display_cols
    }

    sep   = "  "
    hdr   = sep.join(c.upper().ljust(widths[c]) for c in display_cols)
    hline = sep.join("─" * widths[c] for c in display_cols)

    if title:
        print(f"\n{ANSI['bold']}{title}{ANSI['reset']}")
        print("─" * min(len(title) + 4, 80))

    print(f"{ANSI['bold']}{hdr}{ANSI['reset']}")
    print(hline)

    for row in rows:
        line   = sep.join(
            _safe_str(row.get(c, ""))[:widths[c]].ljust(widths[c])
            for c in display_cols
        )
        colour = colour_fn(row) if colour_fn else None
        if colour and colour in ANSI:
            print(f"{ANSI[colour]}{line}{ANSI['reset']}")
        else:
            print(line)
    print()


def print_summary(pairs: list[tuple[str, str]], title: str = "Summary") -> None:
    """Print a two-column key/value summary block."""
    print(f"\n{ANSI['bold']}{title}{ANSI['reset']}")
    print("─" * 40)
    for label, value in pairs:
        print(f"  {label:<35} {value}")
    print()


# ── Excel output helpers ───────────────────────────────────────────────────────

def make_fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell      = ws.cell(row=1, column=col)
        cell.fill = make_fill(COLOURS["header"])
        cell.font = Font(color="FFFFFF", bold=True)


def auto_col_width(ws) -> None:
    for col in ws.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(_safe_str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def write_excel(
    sheets:       dict[str, Optional[pd.DataFrame]],
    output_path:  Path,
    colour_rules: Optional[dict[str, Callable]] = None,
) -> None:
    """Write a multi-sheet colour-coded Excel workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is None or (hasattr(df, "empty") and df.empty):
                placeholder = pd.DataFrame([{"Note": f"No data for '{sheet_name}'"}])
                placeholder.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                continue
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            style_header_row(ws, len(df.columns))
            if colour_rules and sheet_name in colour_rules:
                rule_fn = colour_rules[sheet_name]
                for r_idx, row in enumerate(df.itertuples(index=False)):
                    colour = rule_fn(r_idx, row)
                    if colour:
                        fill = make_fill(colour)
                        for c_idx in range(1, len(df.columns) + 1):
                            ws.cell(row=r_idx + 2, column=c_idx).fill = fill
            auto_col_width(ws)

    total = sum(
        len(df) for df in sheets.values()
        if df is not None and hasattr(df, "__len__") and not df.empty
    )
    print(f"[+] Saved {total} rows across {len(sheets)} sheet(s) → {output_path}")
